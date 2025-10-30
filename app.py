import streamlit as st
import json
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side
import tempfile

# === KONFIGURASI ===
st.set_page_config(page_title="Monitoring Kunjungan Alfamidi", page_icon="store", layout="centered")

# === DATA SEMENTARA ===
DATA_FILE = "data_kunjungan.json"

def load_data():
    if os.path.exists(DATA_FILE):
        try:
            with open(DATA_FILE, "r") as f:
                return json.load(f)
        except:
            return {"toko": [], "selected": [], "petugas": {}}
    return {"toko": [], "selected": [], "petugas": {}}

def save_data(data):
    with open(DATA_FILE, "w") as f:
        json.dump(data, f, indent=4)

# === INIT ===
data = load_data()

# === HEADER ===
st.title("DASHBOARD MONITORING KUNJUNGAN TOKO ALFAMIDI")

# === FORM PETUGAS ===
col1, col2 = st.columns(2)
bulan = col1.selectbox("Pilih Bulan",
    ["Januari", "Februari", "Maret", "April", "Mei", "Juni",
     "Juli", "Agustus", "September", "Oktober", "November", "Desember"],
    index=datetime.now().month - 1)
tahun = col2.number_input("Tahun", min_value=2020, max_value=2100, value=datetime.now().year)

nama = st.text_input("Nama Petugas", value=data["petugas"].get("nama", ""))
nik = st.text_input("NIK", value=data["petugas"].get("nik", ""))
jabatan = st.text_input("Jabatan", value=data["petugas"].get("jabatan", ""))

# Simpan petugas
data["petugas"] = {"nama": nama, "nik": nik, "jabatan": jabatan}
save_data(data)

st.markdown("---")

# === DAFTAR TOKO ===
st.subheader("Pilih Toko yang Dikunjungi")

toko_default = [
    "ALFAMIDI GALESONG TAKALAR", "ALFAMIDI BONTOSUNGGU TAKALAR", "ALFAMIDI GALESONG UTARA TAKALAR",
    "ALFAMIDI BAROMBONG", "ALFAMIDI EMMY SAELAN", "ALFAMIDI TALASALAPANG", "ALFAMIDI JIPANG RAYA",
    "ALFAMIDI MINASA UPA", "ALFAMIDI SULTAN HASANUDDIN", "ALFAMIDI AGUS SALIM", "ALFAMIDI POROS MALINO",
    "ALFAMIDI POROS PALANGGA", "ALFAMIDI METRO TANJUNG BUNGA", "ALFAMIDI METRO TANJUNG BUNGA BLOK H",
    "ALFAMIDI BONTO KADOPEPE", "ALFAMIDI SUPER LIMBUNG", "ALFAMIDI ELANG BANTAENG", "ALFAMIDI RATULANGI BANTAENG",
    "ALFAMIDI RAYA LANTO BANTAENG", "ALFAMIDI PAHLAWAN BANTAENG", "ALFAMIDI BANGKALA JENEPONTO",
    "ALFAMIDI POROS TAKALAR", "ALFAMIDI PALEKO TAKALAR", "ALFAMIDI SAM RATULANGI BULUKUMBA",
    "ALFAMIDI GAJAH MADA BULUKUMBA", "ALFAMIDI PAHLAWAN BULUKUMBA", "ALFAMIDI ANDI SULTAN BULUKUMBA",
    "ALFAMIDI KUSUMA BANGSA BULUKUMBA", "ALFAMIDI JEND. SUDIRMAN BULUKUMBA", "ALFAMIDI BONTOBAHARI BULUKUMBA",
    "ALFAMIDI KALIMPORO KAJANG BULUKUMBA", "ALFAMIDI KEMAKMURAN TANETE BULUKUMBA", "ALFAMIDI PAHLAWAN 2 JENEPONTO",
    "ALFAMIDI PAHLAWAN JENEPONTO", "ALFAMIDI JENEPONTO 2", "ALFAMIDI LANTO PASEWANG JENEPONTO",
    "ALFAMIDI TAMALATEA JENEPONTO", "ALFAMIDI TOLO JENEPONTO", "ALFAMIDI RUMBIA JENEPONTO"
]

if not data["toko"]:
    data["toko"] = toko_default
    save_data(data)

# Tambah toko baru
new_toko = st.text_input("Tambah Toko Baru (opsional)")
if st.button("Tambah Toko"):
    if new_toko and new_toko.strip() and new_toko not in data["toko"]:
        data["toko"].append(new_toko)
        save_data(data)
        st.success(f"Toko '{new_toko}' ditambahkan!")
        st.rerun()

# Pilih toko
cols = st.columns(3)
selected_toko = []
for i, toko in enumerate(data["toko"]):
    if cols[i % 3].checkbox(toko, key=f"chk_{i}", value=(toko in data["selected"])):
        selected_toko.append(toko)

data["selected"] = selected_toko
save_data(data)

st.markdown("---")

# === UPLOAD & SIMPAN KE EXCEL ===
uploaded_file = st.file_uploader("Upload Template Excel Bulanan", type=["xlsx"])

if uploaded_file and st.button("SIMPAN KE EXCEL", type="primary"):
    if not selected_toko:
        st.error("Pilih minimal 1 toko!")
    elif not nama:
        st.error("Nama petugas wajib diisi!")
    else:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp.write(uploaded_file.getvalue())
            file_path = tmp.name

        try:
            wb = load_workbook(file_path)
            ws = wb.active

            # === BORDER PENUH UNTUK SEMUA SEL ===
            full_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            center_align = Alignment(horizontal='center', vertical='center')
            left_align = Alignment(horizontal='left', vertical='center')

            # === 1. ISI HEADER HANYA DI BARIS 4 (SEKALI) ===
            if not ws["B4"].value:
                ws["B4"] = bulan
                ws["C4"] = nama
                ws["D4"] = nik
                ws["E4"] = jabatan

                for col in ["B", "C", "D", "E"]:
                    cell = ws[f"{col}4"]
                    cell.alignment = center_align
                    cell.border = full_border

            # === 2. CARI BARIS KOSONG MULAI DARI BARIS 4 (kolom F) ===
            row = 4
            while ws[f"F{row}"].value is not None:
                row += 1

            # === 3. ISI DATA TOKO (A & F) ===
            for i, toko in enumerate(selected_toko):
                curr_row = row + i
                # Kolom A: No urut
                ws[f"A{curr_row}"] = i + 1
                ws[f"A{curr_row}"].alignment = center_align
                ws[f"A{curr_row}"].border = full_border

                # Kolom F: Nama Toko
                ws[f"F{curr_row}"] = toko
                ws[f"F{curr_row}"].alignment = left_align
                ws[f"F{curr_row}"].border = full_border

            # === 4. TERAPKAN GARIS PENUH KE SEMUA SEL A–F DARI BARIS 4 SAMPAI AKHIR DATA ===
            max_row = row + len(selected_toko) - 1
            for r in range(4, max_row + 1):
                for col_letter in ["A", "B", "C", "D", "E", "F"]:
                    cell = ws[f"{col_letter}{r}"]
                    cell.border = full_border  # Pastikan semua punya garis

            # Simpan
            wb.save(file_path)

            # Download
            with open(file_path, "rb") as f:
                st.download_button(
                    label="DOWNLOAD FILE EXCEL (TABEL PENUH GARIS)",
                    data=f.read(),
                    file_name=f"Kunjungan_{nama}_{bulan}_{tahun}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

            st.success(f"Berhasil! Header hanya di baris 4. {len(selected_toko)} toko ditambahkan. Tabel penuh garis.")
            st.balloons()

        except Exception as e:
            st.error(f"Error: {e}")
        finally:
            os.unlink(file_path)
else:
    st.info("Upload file Excel → Isi data → Pilih toko → Klik *SIMPAN KE EXCEL*")